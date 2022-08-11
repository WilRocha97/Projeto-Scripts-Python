# -*- coding: utf-8 -*-
import time
import openpyxl
from requests import Session
from selenium import webdriver
from selenium.webdriver.support.ui import Select
from bs4 import BeautifulSoup
import os
import re
from datetime import datetime
import sys
sys.path.append("..")
from comum import pfx_to_pem, salvar_arquivo


empresas = [ 
    ['07158353000163', r'..\certificados\CERT RPEM 35586086.pfx', '35586086'],
]

# Verifica se existe alguma mensagem de alerta na tela, retornando False caso encontre
def verifica_alerta(driver):
    try:
        modal = driver.find_element_by_id('dialog-modal')
        print('>>> ', modal.text)
        if modal:
            driver.find_element_by_tag_name('button').click()
            print('>>> Programa Finalizado.')
            return True
    except:
        return False


# Confirma a página acessada, retornando True caso seja diferente do esperado
def verifica_pagina(driver, url):
    if driver.current_url == url:
        print('>>> Página atual: ', driver.current_url)
        return False
    else: 
        print('Erro ao carregar página.', driver.current_url)
        return True


def dados_atualizados(soup):
    viewstate = soup.select("#__VIEWSTATE")[0]['value']
    eventvalidation = soup.select("#__EVENTVALIDATION")[0]['value']

    return (viewstate, eventvalidation)


def analisa_planilha(planilha):
    linha, coluna, total = [0,0,0]
    arquivo = openpyxl.load_workbook(planilha)
    nome_tabela = arquivo.get_sheet_names()[0]
    tabela = arquivo.get_sheet_by_name(nome_tabela)

    for i in range(1, 500):
        if tabela.cell(row=i, column=2).value == 'Total do mês':
            linha = i
            break
    for j in range(1, 20):
        if tabela.cell(row=9, column=j).value == '(5)':
            coluna = j
            break

    if linha and coluna:
        total = float(tabela.cell(row=linha, column=coluna).value)

    return total


def pesquisa_relatorios(cpfcnpj, driver_cookies, competencia, ano):
    url_relatorio = "https://satsp.fazenda.sp.gov.br/COMSAT/Private/ConsultarRelatoriosEscrituracao/ConsultarRelatoriosEscrituracao.aspx"
    regex = re.compile(r"javascript:__doPostBack\('(.+)',''\)")
    saldo_total = 0

    # Inicia uma página para realizar o download dos arquivos
    with Session() as s:
        for cookies in driver_cookies:
            s.cookies.set(cookies['name'], cookies['value'])

        pagina = s.get(url_relatorio, verify=False)
        soup = BeautifulSoup(pagina.content, "html.parser")
        state, validation = dados_atualizados(soup)

        info = {
            'ToolkitScriptManager1_HiddenField': '', '__EVENTTARGET': '',
            '__EVENTARGUMENT': '', '__VIEWSTATE': state, '__VIEWSTATEGENERATOR': 'D6BC1099',
            '__SCROLLPOSITIONX': '0', '__SCROLLPOSITIONY': '0', '__EVENTVALIDATION': validation,
            'ctl00$conteudo$ddlMesReferencia': competencia, 'ctl00$conteudo$ddlAnoReferencia': ano,
            'ctl00$conteudo$btnPesquisar': 'Pesquisar'
        }

        pagina = s.post(url_relatorio, info)
        soup = BeautifulSoup(pagina.content, "html.parser")
        info['__VIEWSTATE'], info['__EVENTVALIDATION'] = dados_atualizados(soup)
        info['__EVENTTARGET'] = 'btnConfirmarClickServer'
        info['__EVENTARGUMENT'] = 'Confirmar'
        info.pop('ctl00$conteudo$btnPesquisar')

        pagina = s.post(url_relatorio, info)
        soup = BeautifulSoup(pagina.content, "html.parser")
        info['__VIEWSTATE'], info['__EVENTVALIDATION'] = dados_atualizados(soup)
        info['__EVENTARGUMENT'] = ''
        info.pop('ctl00$conteudo$ddlMesReferencia')
        info.pop('ctl00$conteudo$ddlAnoReferencia')
        
        try:
            relatorios = soup.find('table', attrs={'id': 'conteudo_grvRelatoriosEscrituracao'}).findAll('a')
        except:
            print('Nenum arquivo encontrado.')
            return False
            
        for index, r in enumerate(relatorios):
            href = r.get('href')
            info['__EVENTTARGET'] = str(regex.search(href).group(1))
            relatorio = s.post(url_relatorio, info)

            filename = relatorio.headers.get('content-disposition', '')

            if filename:
                nome = '_'.join([str(index), cpfcnpj])+'.xlsx'
                salvar_arquivo(nome, relatorio)
                saldo_total += analisa_planilha(os.path.join('documentos', nome))

        print('>>> Saldo total:', saldo_total)

        return True


def consulta_sat(cpfcnpj, certificado, senha, competencia, ano):
    url_login = "https://satsp.fazenda.sp.gov.br/COMSAT/Account/LoginSSL.aspx"
    url_selecionar_cnpj = "https://satsp.fazenda.sp.gov.br/COMSAT/Private/SelecionarCNPJ/SelecionarCNPJContribuinte.aspx"

    cnpj = [cpfcnpj[:8], cpfcnpj[8:]]

    with pfx_to_pem(certificado, senha) as cert:
        driver = webdriver.PhantomJS(os.path.join('..', 'phantomjs.exe'), service_args=['--ssl-client-certificate-file='+cert])
        driver.set_window_size(1000, 900)
        driver.delete_all_cookies()
        
        # Acessa a página inicial
        try:
            driver.get(url_login)
        except Exception as e:
            print(e)
            driver.quit()
            return False
        time.sleep(1)
        driver.find_element_by_id("conteudo_rbtContabilista").click()
        time.sleep(1)
        driver.find_element_by_id("conteudo_imgCertificado").click()

        # Após logar, acessa a página para selecionar um cnpj
        try:
            driver.get(url_selecionar_cnpj)
        except Exception as e:
            print(e)
            driver.quit()
            return False
        if verifica_pagina(driver, url_selecionar_cnpj): 
            driver.quit()
            return False

        # Inclui os dados do cnpj nos campos de busca e pesquisa
        driver.execute_script(f"document.getElementById('conteudo_txtCNPJ_ContribuinteNro').value = '{cnpj[0]}';")
        driver.execute_script(f"document.getElementById('conteudo_txtCNPJ_ContribuinteFilial').value = '{cnpj[1]}';")
        driver.find_element_by_id('conteudo_btnPesquisar').click()
        time.sleep(2)
        if verifica_alerta(driver): 
            driver.quit()
            return True

        # Seleciona o cnpj restante da pesquisa
        driver.find_element_by_id('conteudo_gridCNPJ_lnkCNPJ_0').click()
        time.sleep(1)
        if verifica_alerta(driver): 
            driver.quit()
            return True

        pesquisa_relatorios(cpfcnpj, driver.get_cookies(), competencia, ano)

    return True



if __name__ == '__main__':
    comeco = datetime.now()
    competencia = '8'
    ano = '2020'

    try:
        for empresa in empresas:
            consulta_sat(*empresa, competencia, ano)
    except Exception as e:
        print(e)

    fim = datetime.now()
    print('\n>>> Tempo Total:', (fim-comeco))

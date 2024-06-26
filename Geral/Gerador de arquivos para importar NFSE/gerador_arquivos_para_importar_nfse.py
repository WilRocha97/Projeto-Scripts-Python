# -*- coding: utf-8 -*-
import atexit, sys, fitz, PySimpleGUI as sg
import re

from bs4 import BeautifulSoup
from datetime import datetime
from openpyxl import load_workbook
from os import makedirs, path, startfile, remove, getpid, listdir
from pandas import read_excel
from pyautogui import alert, confirm
from re import compile, search
from requests import Session
from shutil import move
from threading import Thread
from time import sleep
from xlrd import open_workbook


def create_lock_file(lock_file_path):
    try:
        # Tente criar o arquivo de trava
        with open(lock_file_path, 'x') as lock_file:
            lock_file.write(str(getpid()))
        return True
    except FileExistsError:
        # O arquivo de trava já existe, indicando que outra instância está em execução
        return False


def remove_lock_file(lock_file_path):
    try:
        remove(lock_file_path)
    except FileNotFoundError:
        pass


def open_lista_dados(input_excel):
    workbook = ''
    tipo_dados = ''
    file = input_excel
    
    if not file:
        return False
    
    # try:
    # abre se for .xls
    if file.endswith('.xls') or file.endswith('.XLS'):
        workbook = open_workbook(file)
        workbook = workbook.sheet_by_index(0)
        tipo_dados = 'xls'
    
    # abre se for .xlsx
    elif file.endswith('.xlsx') or file.endswith('.XLSX'):
        workbook = load_workbook(file)
        workbook = workbook['Plan1']
        tipo_dados = 'xlsx'
    
    """# abre um alerta se não conseguir abrir o arquivo
    except Exception as e:
        alert(title='Mensagem erro', text=f'Não pode abrir arquivo\n{str(e)}')
        return False"""

    return file, workbook, tipo_dados


# Recebe um texto 'texto' junta com 'end' e escreve num arquivo 'nome'
def escreve_relatorio_csv(texto, nome='resumo', local='', end='\n', encode='latin-1'):
    makedirs(local, exist_ok=True)

    try:
        f = open(path.join(local, f"{nome}.csv"), 'a', encoding=encode)
    except:
        f = open(path.join(local, f"{nome} - auxiliar.csv"), 'a', encoding=encode)

    f.write(texto + end)
    f.close()


def escreve_doc(texto, nome='log', local='Log', encode='latin-1'):
    makedirs(local, exist_ok=True)
    
    try:
        f = open(path.join(local, f"{nome}.txt"), 'a', encoding=encode)
    except:
        f = open(path.join(local, f"{nome} - auxiliar.txt"), 'a', encoding=encode)
    
    f.write(f'{texto}\n\n')
    f.close()

    
def salvar_arquivo(nome, response, pasta_final):
    makedirs(path.join(pasta_final, 'Notas'), exist_ok=True)
    arquivo = open(path.join(pasta_final, 'Notas', nome), 'wb')
    
    # salva o PDF baseado na resposta da requisição com a chave da nota
    for parte in response.iter_content(100000):
        arquivo.write(parte)
    arquivo.close()

    
def download_nota(nota, s, pasta_final):
    chave = str(nota)
    
    # verifica se a chave de acesso não contem letras, se contem anota e retorna para executar a próxima
    if search(r'[a-zA-Z]', chave):
        escreve_relatorio_csv(f"'{nota};Não é uma chave válida", nome='Andamentos', local=pasta_final)
        return False
    
    query = {'codigo': chave,
             'operacao': 'validar'}

    # faz a validação das notas
    response = s.post('https://valinhos.sigissweb.com/nfecentral?oper=validanfe&codigo={}&tipo=V'.format(chave), data=query)
    # define o nome do PDF da nota
    nome_nota = 'nfe_' + chave + '.pdf'
    soup = BeautifulSoup(response.content, 'html.parser', from_encoding="iso-8859-1")
    soup = soup.prettify()
    
    # verifica se o número da chave é uma chave válida, se não, anota e retorna para executar a próxima
    if compile(r'O Código de Autenticidade da Nota Fiscal eletrônica informado é inválido.').search(soup):
        escreve_relatorio_csv(f"'{chave};O Código de Autenticidade da Nota Fiscal eletrônica informado é inválido.", nome='Andamentos', local=pasta_final)
        s.get('https://valinhos.sigissweb.com/validarnfe')
        return False
    
    # salva o PDF da nota
    salvar_arquivo(nome_nota, response, pasta_final)
    sleep(0.5)
    # analisa a nota e coleta informações
    analisa_nota(nome_nota, pasta_final)
    
    return True
    

def analisa_nota(nome_nota, pasta_final):
    # Abrir o pdf
    arq = path.join(pasta_final, 'Notas', nome_nota)
    with fitz.open(arq) as pdf:
        # Para cada página do pdf
        for count, page in enumerate(pdf):
            # Pega o texto da pagina
            textinho = page.get_text('text', flags=1 + 2 + 8)
            
            cancelada = compile(r'CANCELADA').search(textinho)
        
            try:
                cnpj_cliente = compile(r'(\d\d\.\d\d\d\.\d\d\d/\d\d\d\d-\d\d)\nTELEFONE / FAX').search(textinho).group(1)
            except:
                try:
                    cnpj_cliente = compile(r'(\d\d\d\.\d\d\d\.\d\d\d-\d\d)\nTELEFONE / FAX').search(textinho).group(1)
                except:
                    cnpj_cliente = ''
            cnpj_cliente = cnpj_cliente.replace('.', '').replace('/', '').replace('-', '')
            
            nome_cliente = compile(r'DATA DO CADASTRO.+\n(.+)').search(textinho).group(1)
            if nome_cliente == 'IMPOSTOS FEDERAIS RETIDOS':
                nome_cliente = compile(r'DATA DO CADASTRO(.+\n){5}(.+)').search(textinho).group(2)
            
            try:
                uf_cliente = compile(r'UF\n(.+)').search(textinho).group(1)
            except:
                uf_cliente = ''
                
            try:
                municipio_cliente = compile(r'MUNICÍPIO\n.+\n.+\n(.+)').search(textinho).group(1)
            except:
                municipio_cliente = ''
            
            endereco_cliente = compile(r'(.+)\n.+\nTELEFONE / FAX').search(textinho).group(1)
            if endereco_cliente == 'INSCRIÇÃO ESTADUAL':
                endereco_cliente = compile(r'(.+)\nTELEFONE / FAX').search(textinho).group(1)
                
            numero_nota = compile(r'(.+)\nNÚMERO').search(textinho).group(1)
            data_nota = compile(r'(.+)\n.+\nDATA EMISSÃO').search(textinho).group(1)
            
            acumulador = '2102'
                
            valor_servico = compile(r'VALOR BRUTO DA NOTA FISCAL\n.+\n.+\nR\$ (.+)').search(textinho).group(1)
            
            # se o ISS não for retido o valor fica no campo do ISS, se for retido, ele fica no campo do ISS retido
            if compile(r'O ISS NÃO DEVE SER RETIDO'):
                valor_iss = '0,00'
                aliquota_iss = '0,00'
                base_de_calculo = '0,00'
            else:
                valor_iss = compile(r'R\$ (.+)\n(.+\n){5}VALOR I.S.S.').search(textinho).group(1)
                aliquota_iss = compile(r'(\d,\d\d)(.+\n){7}ALIQUOTA ISS\(%\)').search(textinho).group(1)
                base_de_calculo = valor_servico
            
            # tenta pegar o valor dos impostos a seguir, se não conseguir, coloca '0,00' ná variável
            try:
                valor_irrf = compile(r'R\$ (.+)\n.+\nIRRF').search(textinho).group(1)
            except:
                valor_irrf = '0,00'
            try:
                valor_pis = compile(r'R\$ (.+)\n.+\nPIS').search(textinho).group(1)
            except:
                valor_pis = '0,00'
            try:
                valor_cofins = compile(r'R\$ (.+)\n.+\nCOFINS').search(textinho).group(1)
            except:
                valor_cofins = '0,00'
            try:
                valor_csll = compile(r'R\$ (.+)\n.+\nCSLL').search(textinho).group(1)
            except:
                valor_csll = '0,00'
            
            base_de_calculo_inss = '0,00'
            valor_inss = '0,00'
            
            dados_notas_cliente = ';'.join([cnpj_cliente, nome_cliente, uf_cliente, municipio_cliente, endereco_cliente])

            dados_notas = ';'.join([data_nota,
                                    numero_nota,
                                    '26.973.312/0001-75',
                                    'R. POSTAL SERVICOS CONTABEIS LTDA',
                                    'NFS-E;NFD',
                                    '',
                                    valor_servico.replace('.', ''),
                                    base_de_calculo.replace('.', ''),
                                    aliquota_iss.replace('.', ''),
                                    valor_iss.replace('.', ''),
                                    valor_irrf.replace('.', ''),
                                    valor_pis.replace('.', ''),
                                    valor_cofins.replace('.', ''),
                                    valor_csll.replace('.', ''),
                                    base_de_calculo_inss.replace('.', ''),
                                    valor_inss.replace('.', ''),
                                    acumulador,
                                    '1933',
                                    'SP'])
            
            pasta_final_nota = path.join(pasta_final, 'Notas')
            nome_nota_final = f'nfe_{numero_nota}.pdf'
            
            # se for CPF, anota ocorrência, renomeia a nota com o CPF do tomador e move para uma pasta separada
            if 0 < len(cnpj_cliente) < 14:
                escreve_relatorio_csv(f"NFSe_{numero_nota}.pdf;Tomador com CPF, não gera arquivo para importar;'{cnpj_cliente}", nome='Andamentos', local=pasta_final)
                pasta_final_nota = path.join(pasta_final, 'Notas CPF')
                nome_nota_final = f'nfe_{numero_nota}_tomador_{cnpj_cliente}.pdf'
                makedirs(pasta_final_nota, exist_ok=True)
            
            # verifica se a nota está cancelada
            elif cancelada:
                escreve_relatorio_csv(f"NFSe_{numero_nota}.pdf;Nota fiscal cancelada;'{cnpj_cliente}", nome='Andamentos', local=pasta_final)
                nome_nota_final = f'nfe_{numero_nota} - CANCELADA.pdf'
               
            # se não constar CNPJ do cliente na nota, anota a ocorrência
            elif cnpj_cliente == '':
                escreve_relatorio_csv(f"NFSe_{numero_nota}.pdf;Tomador sem CNPJ informado", nome='Andamentos', local=pasta_final)
            
            # se tiver busca o código do Domínio na planilha
            else:
                codigo_cliente = busca_codigo_cliente(cnpj_cliente)
                # se não encontrar o CNPJ do cliente na planilha de códigos, anota a ocorrência
                if codigo_cliente == 'Cliente não encontrado na planilha de códigos':
                    escreve_relatorio_csv(f"NFSe_{numero_nota}.pdf;{codigo_cliente};'{cnpj_cliente}", nome='Andamentos', local=pasta_final)
                elif codigo_cliente == 'CÓDIGO DO DOMÍNIO':
                    escreve_relatorio_csv(f"NFSe_{numero_nota}.pdf;CNPJ do cliente repetido na planilha de códigos, verificar qual consta com o código correto e deletar os demais;'{cnpj_cliente}", nome='Andamentos', local=pasta_final)
                # se o CNPJ do cliente for encontrado na planilha de códigos, cria arquivo .txt para importar no dóminio web
                else:
                    cria_txt(codigo_cliente, cnpj_cliente, dados_notas, pasta_final)
                    escreve_relatorio_csv(f"NFSe_{numero_nota}.pdf;Arquivo para importação criado com sucesso;{codigo_cliente} - {cnpj_cliente}", nome='Andamentos', local=pasta_final)
                    
            escreve_relatorio_csv(f'{dados_notas_cliente};{dados_notas}', nome='Dados das notas', local=pasta_final)

    # renomeia a nota colocando o número dela
    move(arq, path.join(pasta_final_nota, nome_nota_final))
    return


def cria_txt(codigo_cliente, cnpj_cliente, dados_notas, pasta_final, tipo='NFSe', encode='latin-1'):
    # cria um txt com os dados das notas em uma pasta nomeada com o código no domínio e o cnpj do prestador
    local = path.join(pasta_final, 'Arquivos para Importação', str(codigo_cliente) + '-' + str(cnpj_cliente))
    makedirs(local, exist_ok=True)
    
    try:
        f = open(path.join(local, f"{tipo} - {cnpj_cliente}.txt"), 'a', encoding=encode)
    except:
        f = open(path.join(local, f"{tipo}  - {cnpj_cliente} - auxiliar.txt"), 'a', encoding=encode)
    
    f.write(str(dados_notas) + '\n')
    f.close()
    

def busca_codigo_cliente(cnpj_cliente):
    try:
        # Localiza a planilha
        caminho = path.join('Dados_clientes', 'códigos clientes.xlsx')
        # Abrir a planilha
        lista = read_excel(caminho)
        # Definir o index da planilha
        lista.set_index('CNPJ', inplace=True)
        
        cliente = lista.loc[int(cnpj_cliente)]
        cliente = list(cliente)
        codigo_tomador = cliente[0]
        
        return codigo_tomador
    except:
        return 'Cliente não encontrado na planilha de códigos'
    
    
def captura_dados_xml(window_xml, count, numeros_notas, quantidade_notas, arq_name, arq, pasta_final, cidade, numero_da_nota, cnpj_cliente, regex_tomador_cnpj, busca_dados_clientes, busca_dados_notas):
    # print(arq)
    # sleep(55)
    try:
        for i in range(200):
            try:
                cnpj_cliente = compile(numero_da_nota + '(\n.+){' + str(i) + '}' + regex_tomador_cnpj).search(arq).group(2)
                break
            except:
                pass
            
        dados_notas_cliente = []
        for busca_dados in busca_dados_clientes:
            for i in range(500):
                try:
                    dados_notas_cliente.append(compile(r'{}'.format( str(busca_dados[0]) + str(i) + str(busca_dados[1]) )).search(arq).group(2))
                    break
                except:
                    pass
                if i >= 500:
                    escreve_relatorio_csv(f"{arq_name};Existe nota com layout diferente dentro do XML;{numero_da_nota}", nome='Andamentos', local=pasta_final)
                    
        dados_notas_cliente = ';'.join(dados_notas_cliente)
        
        dados_notas = []
        for busca_dados in busca_dados_notas:
            for i in range(500):
                try:
                    dado = compile(r'{}'.format(str(busca_dados[0]) + str(i) + str(busca_dados[1]))).search(arq).group(2)
                    if compile(r'\d\d\d\d-\d\d-\d\dT').search(dado):
                        data_nota_formatada = ''
                        while str(dado[0]) != 'T':
                            data_nota_formatada += dado[0]
                            dado = dado[1:]
                        data_nota = data_nota_formatada.split('-')
                        data_nota = f'{data_nota[2]}/{data_nota[1]}/{data_nota[0]}'
                        dados_notas.append(data_nota)
                    # troca o ponto por vírgula nos valores para importar no sistema
                    elif compile(r'\d.\d').search(dado):
                        dados_notas.append(dado.replace('.', ','))
                    else:
                        dados_notas.append(dado)
                        
                    break
                except:
                    pass
                if i >= 500:
                    escreve_relatorio_csv(f"{arq_name};Existe nota com layout diferente dentro do XML;{numero_da_nota}", nome='Andamentos', local=pasta_final)
            
        while str(numero_da_nota[0]) =='0':
            numero_da_nota = numero_da_nota[1:]
        dados_notas.insert(1, numero_da_nota) #1
        
        if cidade == 'Jundiaí':
            for i in range(500):
                try:
                    cnpj = compile(r'<ns2:Numero>\n\s+' + str(numero_da_nota) + '(\n.+){' + str(i) + '}<ns2:Prestador>\n.\s+<ns2:CpfCnpj>\n.\s+<ns2:Cnpj>\n.\s+(\d+)\n').search(arq).group(2)
                    break
                except:
                    pass
                
            cnpj = re.sub(r'(\d{2})(\d{3})(\d{3})(\d{4})(\d{2})', r'\1.\2.\3/\4-\5', str(cnpj))
            dados_notas.insert(2, cnpj) #2
            
        if cidade == 'Campinas':
            dados_notas.insert(2, '26.973.312/0003-37')  # 2
            
        dados_notas.insert(3, 'R. POSTAL SERVICOS CONTABEIS LTDA') #3
        dados_notas.insert(4, 'NFS-E;NFD') #4 #5
        dados_notas.insert(5, '') #6
    
        dados_notas.insert(7, '0,00') #8
        dados_notas.insert(8, '0,00') #9

        for i in range(500):
            iss_ret = compile(numero_da_nota + r'(\n.+){' + str(i) + '}<VALOR_ISS_RET>\n\s+(.+)\n').search(arq)
            if iss_ret:
                iss_ret = iss_ret.group(2)
                if iss_ret == '0':
                    dados_notas.insert(9, '0,00')  # 10
                else:
                    dados_notas.insert(9, iss_ret)  # 10
                    print(iss_ret)
                break
            
        dados_notas.insert(14, '0,00') #15
        dados_notas.insert(15, '0,00') #16
        dados_notas.insert(16, '2102') #17
        dados_notas.insert(17, '1933') #18
        dados_notas.insert(18, 'SP') #19
        
        dados_notas = ';'.join(dados_notas)
        
        classifica_xml(dados_notas_cliente, dados_notas, cnpj_cliente, arq_name, pasta_final)
        quantidade_notas += 1
            
    except Exception as erro:
        alert(text=f'Erro ao analizar XML, nota número: {numero_da_nota}. Download encerrado.\n\n'
                   'Caso queira reiniciar o download, apague os arquivos gerados anteriormente ou selecione um novo local.\n\n'
                   'O Script não continua uma execução já iniciada.\n\n'
                   'Arquivo Log salvo.\n\n')
        escreve_doc(datetime.now())
        escreve_doc(f'Erro ao analizar XML, nota número: {numero_da_nota}\n{erro}')
        return 'Erro'
    
    window_xml['-progressbar-'].update_bar(count, max=int(len(numeros_notas)))
    window_xml['-Progresso_texto-'].update(str(round(float(count) / int(len(numeros_notas)) * 100, 1)) + '%')
    window_xml.refresh()
    
    return quantidade_notas


def classifica_xml(dados_notas_cliente, dados_notas, cnpj_cliente, arq_name, pasta_final):
    # se for CPF, anota ocorrência, renomeia a nota com o CPF do tomador e move para uma pasta separada
    if 0 < len(cnpj_cliente) < 14:
        escreve_relatorio_csv(f"{arq_name};Tomador com CPF, não gera arquivo para importar;'{cnpj_cliente}", nome='Andamentos', local=pasta_final)
    
    # se não constar CNPJ do cliente na nota, anota a ocorrência
    elif cnpj_cliente == '':
        escreve_relatorio_csv(f"{arq_name};Tomador sem CNPJ informado", nome='Andamentos', local=pasta_final)
    
    # se tiver busca o código do Domínio na planilha
    else:
        codigo_cliente = busca_codigo_cliente(cnpj_cliente)
        # se não encontrar o CNPJ do cliente na planilha de códigos, anota a ocorrência
        if codigo_cliente == 'Cliente não encontrado na planilha de códigos':
            escreve_relatorio_csv(f"{arq_name};{codigo_cliente};'{cnpj_cliente}", nome='Andamentos', local=pasta_final)
        elif codigo_cliente == 'CÓDIGO DO DOMÍNIO':
            escreve_relatorio_csv(f"{arq_name};CNPJ do cliente repetido na planilha de códigos, verificar qual consta com o código correto e deletar os demais;'{cnpj_cliente}", nome='Andamentos', local=pasta_final)
        # se o CNPJ do cliente for encontrado na planilha de códigos, cria arquivo .txt para importar no dóminio web
        else:
            cria_txt(codigo_cliente, cnpj_cliente, dados_notas, pasta_final)
            escreve_relatorio_csv(f"{arq_name};Arquivo para importação criado com sucesso;{codigo_cliente} - {cnpj_cliente}", nome='Andamentos', local=pasta_final)
    
    escreve_relatorio_csv(f'{dados_notas_cliente};{dados_notas}', nome='Dados das notas', local=pasta_final)
        

def run_xml(cidade, window_xml, input_xml, output_dir):
    # configura a pasta final dos arquivos
    quantidade_arquivos = 0
    quantidade_notas = 0
    for count_arquivos, arq_name in enumerate(listdir(input_xml), start=1):
        # Abrir o xml
        arquivo = path.join(input_xml, arq_name)
        window_xml['-Mensagens-'].update(f'Arquivo XML {str(count_arquivos)} / {str(len(listdir(input_xml)))}')
        window_xml.refresh()
        
        with open(arquivo, 'r', encoding='ISO-8859-1') as f:
            data = f.read()
            
            arq_bs = BeautifulSoup(data, "xml")
            arq = arq_bs.prettify()
            
            if cidade == 'Campinas':
                pasta_final = path.join(output_dir, 'Notas Fiscais de Serviço Campinas')
                # regex para pegar todas as notas do arquivo
                regex_numeros_notas = r'<NUM_NOTA>\n\s+(.+)\n'
                # regex para pegar o cnpj do tomador
                regex_tomador_cnpj = '<TOMADOR_CPF_CNPJ>\n\s+(.+)\n'
                # procura todas as notas no arquivo
                numeros_notas = compile(regex_numeros_notas).findall(arq)
                cnpj_cliente = ''
                for count, numero_da_nota in enumerate(numeros_notas, start=1):
                    # lista de regex para os dados do tomador
                    busca_dados_clientes = [[numero_da_nota + '(\n.+){', '}<TOMADOR_CPF_CNPJ>\n\s+(.+)\n'],
                                            [numero_da_nota + '(\n.+){', '}<TOMADOR_RAZAO_SOCIAL>\n\s+(.+)\n'],
                                            [numero_da_nota + '(\n.+){', '}<TOMADOR_UF>\n\s+(.+)\n'],
                                            [numero_da_nota + '(\n.+){', '}<TOMADOR_CIDADE>\n\s+(.+)\n'],
                                            [numero_da_nota + '(\n.+){', '}<TOMADOR_LOGRADOURO>\n\s+(.+)\n']]
                    # lista de regex dos dados da nota
                    busca_dados_notas = [[numero_da_nota + '(\n.+){', '}<DATA_HORA_EMISSAO>\n\s+(\d\d/\d\d/\d\d\d\d)'],  # 0
                                         [numero_da_nota + '(\n.+){', '}<VALOR_SERVICO>\n\s+(.+)\n'],  # 7
                                         [numero_da_nota + '(\n.+){', '}<VALOR_IR>\n\s+(.+)\n'],  # 11
                                         [numero_da_nota + '(\n.+){', '}<VALOR_PIS>\n\s+(.+)\n'],  # 12
                                         [numero_da_nota + '(\n.+){', '}<VALOR_COFINS>\n\s+(.+)\n'],  # 13
                                         [numero_da_nota + '(\n.+){', '}<VALOR_CSLL>\n\s+(.+)\n']]  # 14
                    
                    # captura os dados na nota e cria os arquivos para importar no domínio
                    quantidade_notas = captura_dados_xml(window_xml, count, numeros_notas, quantidade_notas, arq_name, arq, pasta_final, cidade,
                                                         numero_da_nota, cnpj_cliente, regex_tomador_cnpj, busca_dados_clientes, busca_dados_notas)
                    
                    if quantidade_notas == 'erro':
                        return
                    
                    # Verifica se o usuário solicitou o encerramento do script
                    if event == '-Encerrar-' or event == sg.WIN_CLOSED:
                        break
                    
            if cidade == 'Jundiaí':
                pasta_final = path.join(output_dir, 'Notas Fiscais de Serviço Jundiaí')
                # regex para pegar todas as notas do arquivo
                regex_numeros_notas = r'InfNfse Id=\"\d+\">\n\s+<ns2:Numero>\n\s+(.+)'
                # regex para pegar o cnpj do tomador
                regex_tomador_cnpj = '<ns2:IdentificacaoTomador>\n.+\n.+\n\s+(.+)'
                # procura todas as notas no arquivo
                numeros_notas = compile(regex_numeros_notas).findall(arq)
                cnpj_cliente = ''
                for count, numero_da_nota in enumerate(numeros_notas, start=1):
                    # lista de regex para os dados do tomador
                    busca_dados_clientes = [['<ns2:Numero>\n\s+' + numero_da_nota + '(\n.+){', '}<ns2:IdentificacaoTomador>\n.+\n.+\n\s+(.+)'],
                                            ['<ns2:Numero>\n\s+' + numero_da_nota + '(\n.+){', '}</ns2:IdentificacaoTomador>\n\s+<ns2:RazaoSocial>\n\s+(.+)'],
                                            ['<ns2:Numero>\n\s+' + numero_da_nota + '(\n.+){', '}</ns2:IdentificacaoTomador>\n.+\n.+\n.+\n.+\n.+\n.+\n.+\n.+\n.+\n.+\n.+\n.+\n.+\n.+\n.+\n.+\n.+<ns2:Uf>\n\s+(.+)'],
                                            ['<ns2:Numero>\n\s+' + numero_da_nota + '(\n.+){', '}</ns2:IdentificacaoTomador>\n.+\n.+\n.+\n.+\n.+\n.+\n.+\n.+\n.+\n.+\n.+\n.+\n.+\n.+<ns2:CodigoMunicipio>\n\s+(.+)'],
                                            ['<ns2:Numero>\n\s+' + numero_da_nota + '(\n.+){', '}</ns2:IdentificacaoTomador>\n.+\n.+\n.+\n.+<ns2:Endereco>\n\s+<ns2:Endereco>\n\s+(.+)']]
                    # lista de regex dos dados da nota
                    busca_dados_notas = [['<ns2:Numero>\n\s+' + numero_da_nota + '(\n.+){', '}<ns2:DataEmissao>\n\s+(.+)'],  # 0
                                         ['<ns2:Numero>\n\s+' + numero_da_nota + '(\n.+){', '}<ns2:ValorServicos>\n\s+(.+)\n'],  # 6
                                         ['<ns2:Numero>\n\s+' + numero_da_nota + '(\n.+){', '}<ns2:ValorIr>\n\s+(.+)\n'],  # 10
                                         ['<ns2:Numero>\n\s+' + numero_da_nota + '(\n.+){', '}<ns2:ValorPis>\n\s+(.+)\n'],  # 11
                                         ['<ns2:Numero>\n\s+' + numero_da_nota + '(\n.+){', '}<ns2:ValorCofins>\n\s+(.+)\n'],  # 12
                                         ['<ns2:Numero>\n\s+' + numero_da_nota + '(\n.+){', '}<ns2:ValorCsll>\n\s+(.+)\n']]  # 13
                    # captura os dados na nota e cria os arquivos para importar no domínio
                    quantidade_notas = captura_dados_xml(window_xml, count, numeros_notas, quantidade_notas, arq_name, arq, pasta_final, cidade,
                                                         numero_da_nota, cnpj_cliente, regex_tomador_cnpj, busca_dados_clientes, busca_dados_notas)
                    
                    if quantidade_notas == 'erro':
                        return
                        
                    # Verifica se o usuário solicitou o encerramento do script
                    if event == '-Encerrar-' or event == sg.WIN_CLOSED:
                        break
                    
            quantidade_arquivos += 1
            
        # Verifica se o usuário solicitou o encerramento do script
        if event == '-Encerrar-' or event == sg.WIN_CLOSED:
            alert(text='Download encerrado.\n\n'
                       'Caso queira reiniciar o download, apague os arquivos gerados anteriormente ou selecione um novo local.\n\n'
                       'O Script não continua uma execução já iniciada.\n\n')
            return
            
        window_xml['-progressbar-'].update_bar(count_arquivos, max=int(len(listdir(input_xml))))
        window_xml['-Progresso_texto-'].update(str(round(float(count_arquivos) / int(len(listdir(input_xml))) * 100, 1)) + '%')
        window_xml.refresh()
        
    alert(text=f'Analise concluída, {quantidade_arquivos} arquivos analisados\n'
               f'{quantidade_notas} notas capturadas')
    

def run_valinhos(window_valinhos, input_excel, output_dir):
    arquivo, notas, tipo_dados = open_lista_dados(input_excel)
    # Abre o site do SIGISS pronto para validar as notas
    with Session() as s:
        soup = ''
        timer = 0
        timer_2 = -1
        # enquanto o site não carregar, tenta acessá-lo e caso demore muito retorna
        while not compile(r'<input class=\"\" id=\"codigo\" label=\"Informe o Código de Autenticidade : \" maxlength=\"42\" name=\"codigo\" requerido=\"true\" type=\"text\"/>').search(soup):
            sleep(0.99)
            try:
                response = s.get('https://valinhos.sigissweb.com/validarnfe')
                soup = BeautifulSoup(response.content, 'html.parser', from_encoding="Latin-1")
                soup = soup.prettify()
            except:
                soup = ''
                pass
            
            timer += 1
            if timer >= 2:
                timer_2 += 1
                # exibe um contador de espera na interface
                window_valinhos['-Mensagens-'].update('Aguardando SIGISSWEB... ' + str(timer_2))
                
            if timer >= 32:
                window_valinhos['-Mensagens-'].update('')
                alert(text='SIGISSWEB demorou muito para responder, download encerrado.\n\n'
                           'O site pode estar fora do ar ou foi movido para outro endereço.\n\n'
                           'Verifique se o site "https://valinhos.sigissweb.com/validarnfe" ainda existe.\n\n'
                           'Verifique sua conexão com a internet.\n\n')
                s.close()
                return
                
            # Verifica se o usuário solicitou o encerramento do script
            if event == '-Encerrar-' or event == sg.WIN_CLOSED:
                s.close()
                return
        
        # quando o site abrir
        # configura a pasta final dos arquivos
        pasta_final = path.join(output_dir, 'Notas Fiscais de Serviço Valinhos')
        # inicia a variável do contador de notas baixadas
        quantidade_notas = 0
        
        if tipo_dados == 'xls':
            # cria o indice para cada empresa da lista de dados
            total_empresas = range(notas.nrows)
            for count, nota in enumerate(range(notas.nrows), start=1):
                # pega a última coluna da planilha
                nota = notas.cell_value(nota, -1)
                if download_nota(nota, s, pasta_final):
                    quantidade_notas += 1
                # atualiza a barra de progresso
                window_valinhos['-progressbar-'].update_bar(count, max=int(len(total_empresas)))
                window_valinhos['-Progresso_texto-'].update(str( round( float(count) / int(len(total_empresas)) *100, 1 ) ) + '%')
                window_valinhos.refresh()
                
                # Verifica se o usuário solicitou o encerramento do script
                if event == '-Encerrar-' or event == sg.WIN_CLOSED:
                    alert(text='Download encerrado.\n\n'
                               'Caso queira reiniciar o download, apague os arquivos gerados anteriormente ou selecione um novo local.\n\n'
                               'O Script não continua uma execução já iniciada.\n\n')
                    s.close()
                    return
                
        elif tipo_dados == 'xlsx':
            last_column = notas.max_column
            # cria o indice para cada empresa da lista de dados
            total_empresas = []
            for nota in notas.iter_rows(min_row=1, min_col=last_column, max_col=last_column):
                total_empresas.append(nota)
            
            for count, nota in enumerate(notas.iter_rows(min_row=1, min_col=last_column, max_col=last_column), start=1):
                for chave in nota:
                    if download_nota(chave.value, s, pasta_final):
                        quantidade_notas += 1
                    # atualiza a barra de progresso
                    window_valinhos['-progressbar-'].update_bar(count, max=int(len(total_empresas)))
                    window_valinhos['-Progresso_texto-'].update(str( round( float(count) / int(len(total_empresas)) *100, 1 ) ) + '%')
                    window_valinhos.refresh()
                    
                    # Verifica se o usuário solicitou o encerramento do script
                    if event == '-Encerrar-' or event == sg.WIN_CLOSED:
                        alert(text='Download encerrado.\n\n'
                                   'Caso queira reiniciar o download, apague os arquivos gerados anteriormente ou selecione um novo local.\n\n'
                                   'O Script não continua uma execução já iniciada.\n\n')
                        s.close()
                        return
    
        s.close()
        alert(text=f'Download concluído, {quantidade_notas} notas salvas')
    

# Define o ícone global da aplicação
sg.set_global_icon('Assets/auto-flash.ico')
if __name__ == '__main__':
    # Especifique o caminho para o arquivo de trava
    lock_file_path = 'download_nfse_do_escritorio.lock'
    
    # Verifique se outra instância está em execução
    if not create_lock_file(lock_file_path):
        alert(text="O programa já está em execução.")
        sys.exit(1)
    
    # Defina uma função para remover o arquivo de trava ao final da execução
    atexit.register(remove_lock_file, lock_file_path)

    sg.theme('GrayGrayGray')  # Define o tema do PySimpleGUI
    # sg.theme_previewer()
    # Layout da janela
    
    def janela_menu(): # layout da janela do menu principal
        layout_menu = [
            [sg.Button('Ajuda', border_width=0), sg.Button('Lista de códigos do Domínio', key='-lista_dominio-', border_width=0), sg.Button('Log do sistema', border_width=0, disabled=True)],
            [sg.Text('')],
            [sg.Text('Selecione o local do prestador de serviço das notas', justification='center')],
            [sg.Button('Campinas', key='-campinas-', size=20, border_width=1),
             sg.Button('Jundiaí', key='-jundiai-', size=20, border_width=1),
             sg.Button('Valinhos', key='-valinhos-', size=20, border_width=1)]
        ]
        
        # guarda a janela na variável para manipula-la
        return sg.Window('Gerador de arquivos para importar NFSE', layout_menu, finalize=True)
    
    
    def janela_xml(cidade): # layout da janela dos scripts que analisam os xml
        layout_valinhos = [
            [sg.Button('Ajuda', border_width=0), sg.Button('Lista de códigos do Domínio', key='-lista_dominio-', border_width=0), sg.Button('Log do sistema', border_width=0, disabled=True)],
            [sg.Text('')],
            [sg.Text('Selecione um diretório com os arquivos XML:')],
            [sg.FolderBrowse('Pesquisar', key='-Abrir-'), sg.InputText(key='-input_xml-', size=80, disabled=True)],
            [sg.Text('Selecione um diretório para salvar os resultados:')],
            [sg.FolderBrowse('Pesquisar', key='-Abrir2-'), sg.InputText(key='-output_dir-', size=80, disabled=True)],
            [sg.Text('')],
            [sg.Text('', key='-Mensagens-')],
            [sg.Text(size=6, text='', key='-Progresso_texto-'), sg.ProgressBar(max_value=0, orientation='h', size=(54, 5), key='-progressbar-', bar_color='#f0f0f0')],
            [sg.Button('Iniciar', key='-Iniciar-', border_width=0), sg.Button('Encerrar', key='-Encerrar-', disabled=True, border_width=0), sg.Button('Abrir resultados', key='-Abrir resultados-', disabled=True, border_width=0)],
        ]
        
        # guarda a janela na variável para manipula-la
        return sg.Window(cidade, layout_valinhos, finalize=True, modal=True)
    
    
    def janela_valinhos(): # layout da janela do script das notas de Valinhos
        layout_valinhos = [
            [sg.Button('Ajuda', border_width=0), sg.Button('Lista de códigos do Domínio', key='-lista_dominio-', border_width=0), sg.Button('Log do sistema', border_width=0, disabled=True)],
            [sg.Text('')],
            [sg.Text('Selecione um arquivo Excel com as chaves de acesso das notas:')],
            [sg.FileBrowse('Pesquisar', key='-Abrir-', file_types=(('Planilhas Excel', '*.xlsx *.xls'),)), sg.InputText(key='-input_excel-', size=80, disabled=True)],
            [sg.Text('Selecione um diretório para salvar os resultados:')],
            [sg.FolderBrowse('Pesquisar', key='-Abrir2-'), sg.InputText(key='-output_dir-', size=80, disabled=True)],
            [sg.Text('')],
            [sg.Text('', key='-Mensagens-')],
            [sg.Text(size=6, text='', key='-Progresso_texto-'), sg.ProgressBar(max_value=0, orientation='h', size=(54, 5), key='-progressbar-', bar_color='#f0f0f0')],
            [sg.Button('Iniciar', key='-Iniciar-', border_width=0), sg.Button('Encerrar', key='-Encerrar-', disabled=True, border_width=0), sg.Button('Abrir resultados', key='-Abrir resultados-', disabled=True, border_width=0)],
        ]
        
        # guarda a janela na variável para manipula-la
        return sg.Window('Valinhos', layout_valinhos, finalize=True, modal=True)
    
    # inicia as variáveis das janelas
    window_menu, window_xml, window_valinhos = janela_menu(), None, None
    
    while True:
        # captura o evento e os valores armazenados na interface
        window, event, values = sg.read_all_windows()
        
        if event == sg.WIN_CLOSED:
            if window == window_valinhos:  # if closing win 2, mark as closed
                window_valinhos = None
            if window == window_xml:  # if closing win 2, mark as closed
                window_xml = None
            elif window == window_menu:  # if closing win 1, exit program
                break
        
        elif event == 'Ajuda':
            startfile('Manual do usuário - Gerador de arquivos para importar NFSE.pdf')
        
        elif event == '-lista_dominio-':
            startfile(path.join('Dados_clientes', 'códigos clientes.xlsx'))
            
        elif event == 'Log do sistema':
            startfile('Log')
        
        # se o botão clicado for de campinas ou jundiaí, usa o mesmo layout, porém com o nome específico da cidade
        elif event == '-campinas-' or event == '-jundiai-':
            cidade = ''
            if event == '-campinas-':
                cidade = 'Campinas'
            if event == '-jundiai-':
                cidade = 'Jundiaí'
                
            window_xml = janela_xml(cidade)
            def run_script_thread():
                try:
                    if not input_xml or not output_dir:
                        alert(text=f'Por favor selecione os dois diretórios.')
                        return
                    
                    # habilita e desabilita os botões conforme necessário
                    window_xml['-lista_dominio-'].update(disabled=True)
                    window_xml['-Abrir-'].update(disabled=True)
                    window_xml['-Abrir2-'].update(disabled=True)
                    window_xml['-Iniciar-'].update(disabled=True)
                    window_xml['-Encerrar-'].update(disabled=False)
                    window_xml['-Abrir resultados-'].update(disabled=False)
                    # apaga qualquer mensagem na interface
                    window_xml['-Mensagens-'].update('')
                    # atualiza a barra de progresso para ela ficar mais visível
                    window_xml['-progressbar-'].update(bar_color=('#fca400', '#ffe0a6'))
                    
                    try:
                        # Chama a função que executa o script
                        # nesse caso o script de analisa xml pode ser usado por campinas e jundiaí
                        run_xml(cidade, window_xml, input_xml, output_dir)
                        # Qualquer erro o script exibe um alerta e salva gera o arquivo log de erro
                    except Exception as erro:
                        window_xml['Log do sistema'].update(disabled=False)
                        alert(text=f"Erro :'(\n\n"
                                   f'Abra o pasta de "Log do sistema" e envie o arquivo "Log.txt" para o desenvolvedor.\n')
                        
                        for arq in listdir('Log'):
                            remove(path.join('Log', arq))
                            
                        escreve_doc(datetime.now())
                        escreve_doc(erro)
        
                    # habilita e desabilita os botões conforme necessário
                    window_xml['-lista_dominio-'].update(disabled=False)
                    window_xml['-Abrir-'].update(disabled=False)
                    window_xml['-Abrir2-'].update(disabled=False)
                    window_xml['-Iniciar-'].update(disabled=False)
                    window_xml['-Encerrar-'].update(disabled=True)
                    # apaga a porcentagem e a barra de progresso para a interface ficar mais limpa
                    window_xml['-progressbar-'].update_bar(0)
                    window_xml['-progressbar-'].update(bar_color='#f0f0f0')
                    window_xml['-Progresso_texto-'].update('')
                    window_xml['-Mensagens-'].update('')
                except:
                    pass
            
            while True:
                # captura o evento e os valores armazenados na interface
                event, values = window_xml.read()
                try:
                    input_xml = values['-input_xml-']
                    output_dir = values['-output_dir-']
                except:
                    input_xml = 'Desktop'
                    output_dir = 'Desktop'
                
                if event == sg.WIN_CLOSED:
                    break
                
                elif event == 'Log do sistema':
                    startfile('Log')
                    
                elif event == 'Ajuda':
                    startfile('Manual do usuário - Gerador de arquivos para importar NFSE.pdf')
                  
                elif event == '-lista_dominio-':
                    startfile(path.join('Dados_clientes', 'códigos clientes.xlsx'))
                    
                elif event == '-Iniciar-':
                    # Cria uma nova thread para executar o script
                    script_thread = Thread(target=run_script_thread)
                    script_thread.start()
                    
                elif event == '-Abrir resultados-':
                    pasta_final = ''
                    if cidade == 'Campinas':
                        pasta_final = path.join(output_dir, 'Notas Fiscais de Serviço Campinas')
                    if cidade == 'Jundiaí':
                        pasta_final = path.join(output_dir, 'Notas Fiscais de Serviço Jundiaí')
                        
                    startfile(pasta_final)
                
            window_xml.close()
            
        elif event == '-valinhos-':
            window_valinhos = janela_valinhos()
            def run_script_thread():
                try:
                    if not input_excel or not output_dir:
                        alert(text=f'Por favor selecione uma planilha e um diretório.')
                        return
                    
                    # habilita e desabilita os botões conforme necessário
                    window_valinhos['-lista_dominio-'].update(disabled=True)
                    window_valinhos['-Abrir-'].update(disabled=True)
                    window_valinhos['-Abrir2-'].update(disabled=True)
                    window_valinhos['-Iniciar-'].update(disabled=True)
                    window_valinhos['-Encerrar-'].update(disabled=False)
                    window_valinhos['-Abrir resultados-'].update(disabled=False)
                    # apaga qualquer mensagem na interface
                    window_valinhos['-Mensagens-'].update('')
                    # atualiza a barra de progresso para ela ficar mais visível
                    window_valinhos['-progressbar-'].update(bar_color=('#fca400', '#ffe0a6'))
                    
                    try:
                        # Chama a função que executa o script
                        run_valinhos(window_valinhos, input_excel, output_dir)
                    # Qualquer erro o script exibe um alerta e salva gera o arquivo log de erro
                    except Exception as erro:
                        window_valinhos['Log do sistema'].update(disabled=False)
                        alert(text=f"Erro :'(\n\n"
                                   f'Abra o pasta de "Log do sistema" e envie o arquivo "Log.txt" para o desenvolvedor.\n')
                        
                        for arq in listdir('Log'):
                            remove(path.join('Log', arq))
                            
                        escreve_doc(datetime.now())
                        escreve_doc(erro)
        
                    # habilita e desabilita os botões conforme necessário
                    window_valinhos['-lista_dominio-'].update(disabled=False)
                    window_valinhos['-Abrir-'].update(disabled=False)
                    window_valinhos['-Abrir2-'].update(disabled=False)
                    window_valinhos['-Iniciar-'].update(disabled=False)
                    window_valinhos['-Encerrar-'].update(disabled=True)
                    # apaga a porcentagem e a barra de progresso para a interface ficar mais limpa
                    window_valinhos['-progressbar-'].update_bar(0)
                    window_valinhos['-progressbar-'].update(bar_color='#f0f0f0')
                    window_valinhos['-Progresso_texto-'].update('')
                    window_valinhos['-Mensagens-'].update('')
                except:
                    pass
            
            while True:
                # captura o evento e os valores armazenados na interface
                event, values = window_valinhos.read()
                try:
                    input_excel = values['-input_excel-']
                    output_dir = values['-output_dir-']
                except:
                    input_excel = 'Desktop'
                    output_dir = 'Desktop'
                
                if event == sg.WIN_CLOSED:
                    break
                
                elif event == 'Log do sistema':
                    startfile('Log')
                    
                elif event == 'Ajuda':
                    startfile('Manual do usuário - Gerador de arquivos para importar NFSE.pdf')
                  
                elif event == '-lista_dominio-':
                    startfile(path.join('Dados_clientes', 'códigos clientes.xlsx'))
                    
                elif event == '-Iniciar-':
                    # Cria uma nova thread para executar o script
                    script_thread = Thread(target=run_script_thread)
                    script_thread.start()
                    
                elif event == '-Abrir resultados-':
                    startfile(path.join(output_dir, 'Notas Fiscais de Serviço Valinhos'))
                
            window_valinhos.close()
        
    window_menu.close()
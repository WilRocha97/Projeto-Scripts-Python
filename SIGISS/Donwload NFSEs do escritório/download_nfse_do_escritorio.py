# -*- coding: utf-8 -*-
import PySimpleGUI as sg
from pandas import read_excel
from threading import Thread
from openpyxl import load_workbook
from xlrd import open_workbook
from shutil import move
from os import makedirs, path, startfile
from re import compile, search
from fitz import open as abrir_pdf
from time import sleep
from datetime import datetime
from bs4 import BeautifulSoup
from requests import Session
from pyautogui import alert, confirm


def open_lista_dados(input_excel):
    file = input_excel
    
    if not file:
        return False
    
    try:
        # abre se for .xls
        if file.endswith('.xls') or file.endswith('.XLS'):
            workbook = xlrd.open_workbook(file)
            workbook = workbook.sheet_by_index(0)
            tipo_dados = 'xls'
        
        # abre se for .xlsx
        elif file.endswith('.xlsx') or file.endswith('.XLSX'):
            workbook = load_workbook(file)
            workbook = workbook['Plan1']
            tipo_dados = 'xlsx'
    
    # abre um alerta se não conseguir abrir o arquivo
    except Exception as e:
        alert(title='Mensagem erro', text=f'Não pode abrir arquivo\n{str(e)}')
        return False

    return file, workbook, tipo_dados


# Recebe um texto 'texto' junta com 'end' e escreve num arquivo 'nome'
def escreve_relatorio_csv(texto, nome='resumo', local='', end='\n', encode='latin-1'):
    makedirs(local, exist_ok=True)

    try:
        f = open(path.join(local, f"{nome}.csv"), 'a', encoding=encode)
    except:
        f = open(path.join(local, f"{nome} - auxiliar.csv"), 'a', encoding=encode)

    f.write(texto + end)
    f.close()


def escreve_doc(texto, nome='log', local='Log', encode='latin-1'):
    makedirs(local, exist_ok=True)
    
    try:
        f = open(path.join(local, f"{nome}.txt"), 'a', encoding=encode)
    except:
        f = open(path.join(local, f"{nome} - auxiliar.txt"), 'a', encoding=encode)
    
    f.write(f'--------------------------------------------------------------------------------------------------¬\n{datetime.now()}\n\n'
            f'{texto}\n\n\n')
    f.close()

    
def salvar_arquivo(nome, response, pasta_final):
    makedirs(path.join(pasta_final, 'Notas'), exist_ok=True)
    arquivo = open(path.join(pasta_final, 'Notas', nome), 'wb')
    
    # salva o PDF baseado na resposta da requisição com a chave da nota
    for parte in response.iter_content(100000):
        arquivo.write(parte)
    arquivo.close()

            
def download_nota(nota, s, pasta_final):
    chave = str(nota)
    
    # verifica se a chave de acesso não contem letras, se contem anota e retorna para executar a próxima
    if search(r'[a-zA-Z]', chave):
        escreve_relatorio_csv(f"'{nota};Não é uma chave válida", nome='Andamentos', local=pasta_final)
        return False
    
    query = {'codigo': chave,
             'operacao': 'validar'}

    # faz a validação das notas
    response = s.post('https://valinhos.sigissweb.com/nfecentral?oper=validanfe&codigo={}&tipo=V'.format(chave), data=query)
    # define o nome do PDF da nota
    nome_nota = 'nfe_' + chave + '.pdf'
    soup = BeautifulSoup(response.content, 'html.parser', from_encoding="iso-8859-1")
    soup = soup.prettify()
    
    # verifica se o número da chave é uma chave válida, se não, anota e retorna para executar a próxima
    if compile(r'O Código de Autenticidade da Nota Fiscal eletrônica informado é inválido.').search(soup):
        escreve_relatorio_csv(f"'{chave};O Código de Autenticidade da Nota Fiscal eletrônica informado é inválido.", nome='Andamentos', local=pasta_final)
        s.get('https://valinhos.sigissweb.com/validarnfe')
        return False
    
    # salva o PDF da nota
    salvar_arquivo(nome_nota, response, pasta_final)
    sleep(0.5)
    # analisa a nota e coleta informações
    analisa_nota(nome_nota, pasta_final)
    
    return True
        

def analisa_nota(nome_nota, pasta_final):
    # Abrir o pdf
    arq = path.join(pasta_final, 'Notas', nome_nota)
    with abrir_pdf(arq) as pdf:
        # Para cada página do pdf
        for count, page in enumerate(pdf):
            # Pega o texto da pagina
            textinho = page.get_text('text', flags=1 + 2 + 8)
            # print(textinho)
            # sleep(12)
            try:
                cnpj_cliente = compile(r'(\d\d\.\d\d\d\.\d\d\d/\d\d\d\d-\d\d)\nTELEFONE / FAX').search(textinho).group(1)
            except:
                try:
                    cnpj_cliente = compile(r'(\d\d\d\.\d\d\d\.\d\d\d-\d\d)\nTELEFONE / FAX').search(textinho).group(1)
                except:
                    cnpj_cliente = ''
            cnpj_cliente = cnpj_cliente.replace('.', '').replace('/', '').replace('-', '')
            
            nome_cliente = compile(r'DATA DO CADASTRO.+\n(.+)').search(textinho).group(1)
            if nome_cliente == 'IMPOSTOS FEDERAIS RETIDOS':
                nome_cliente = compile(r'DATA DO CADASTRO(.+\n){5}(.+)').search(textinho).group(2)
            
            try:
                uf_cliente = compile(r'UF\n(.+)').search(textinho).group(1)
            except:
                uf_cliente = ''
                
            try:
                municipio_cliente = compile(r'MUNICÍPIO\n.+\n.+\n(.+)').search(textinho).group(1)
            except:
                municipio_cliente = ''
            
            endereco_cliente = compile(r'(.+)\n.+\nTELEFONE / FAX').search(textinho).group(1)
            if endereco_cliente == 'INSCRIÇÃO ESTADUAL':
                endereco_cliente = compile(r'(.+)\nTELEFONE / FAX').search(textinho).group(1)
                
            numero_nota = compile(r'(.+)\nNÚMERO').search(textinho).group(1)
            serie_nota = compile(r'ELETRÔNICA DE\nSERVIÇO\n.+\n(.+)').search(textinho).group(1)
            data_nota = compile(r'(.+)\n.+\nDATA EMISSÃO').search(textinho).group(1)
            
            # se a nota for válida a situação é 0 se for cancelada é 2
            situacao = compile(r'DOCUMENTO VALIDADO COM SUCESSO').search(textinho)
            if situacao:
                situacao = '0'
            else:
                situacao = '2'
                
            acumulador = '6102'
            
            # se o município do tomador for valinhos, o CFPS é 9101, se não é 9102
            if municipio_cliente == 'Valinhos':
                cfps = '9101'
            else:
                cfps = '9102'
                
            valor_servico = compile(r'VALOR BRUTO DA NOTA FISCAL\n.+\n.+\nR\$ (.+)').search(textinho).group(1)
            valor_descontos = '0,00'
            valor_deducoes = compile(r'R\$ (.+)\n.+\n.+\n.+\nDEDUÇÕES').search(textinho).group(1)
            valor_contabil = valor_servico
            base_de_calculo = valor_servico
            aliquota_iss = compile(r'(\d,\d\d)(.+\n){7}ALIQUOTA ISS\(%\)').search(textinho).group(1)
            
            # se o ISS não for retido o valor fica no campo do ISS, se for retido, ele fica no campo do ISS retido
            if compile(r'O ISS NÃO DEVE SER RETIDO'):
                valor_iss = compile(r'R\$ (.+)\n(.+\n){5}VALOR I.S.S.').search(textinho).group(1)
                iss_retido = '0,00'
            else:
                valor_iss = '0,00'
                iss_retido = compile(r'R\$ (.+)\n(.+\n){5}VALOR I.S.S.').search(textinho).group(1)
            
            # tenta pegar o valor dos impostos a seguir, se não conseguir, coloca '0,00' ná variável
            try:
                valor_irrf = compile(r'R\$ (.+)\n.+\nIRRF').search(textinho).group(1)
            except:
                valor_irrf = '0,00'
            try:
                valor_pis = compile(r'R\$ (.+)\n.+\nPIS').search(textinho).group(1)
            except:
                valor_pis = '0,00'
            try:
                valor_cofins = compile(r'R\$ (.+)\n.+\nCOFINS').search(textinho).group(1)
            except:
                valor_cofins = '0,00'
            try:
                valor_csll = compile(r'R\$ (.+)\n.+\nCSLL').search(textinho).group(1)
            except:
                valor_csll = '0,00'
            
            valor_crf = '0,00'
            valor_inss = '0,00'
            codigo_do_iten = '17.18'
            quantidade = '1,00'
            valor_unitario = valor_servico
            
            dados_notas_cliente = ';'.join([cnpj_cliente, nome_cliente, uf_cliente, municipio_cliente, endereco_cliente])
            dados_notas_escritorio = ';'.join(['26.973.312/0001-75',
                                    'R. POSTAL SERVICOS CONTABEIS LTDA',
                                    'SP',
                                    'Valinhos',
                                    'Rua Fioravante Basílio Maglio, nº 258',])
            dados_notas = ';'.join([numero_nota,
                                    serie_nota,
                                    data_nota,
                                    situacao,
                                    acumulador,
                                    cfps,
                                    valor_servico,
                                    valor_descontos,
                                    valor_deducoes,
                                    valor_contabil,
                                    base_de_calculo,
                                    aliquota_iss,
                                    valor_iss,
                                    iss_retido,
                                    valor_irrf,
                                    valor_pis,
                                    valor_cofins,
                                    valor_csll,
                                    valor_crf,
                                    valor_inss,
                                    codigo_do_iten,
                                    quantidade,
                                    valor_unitario])
            
            pasta_final_nota = path.join(pasta_final, 'Notas')
            nome_nota_final = f'nfe_{numero_nota}.pdf'
            
            # se for CPF, anota ocorrência, renomeia a nota com o CPF do tomador e move para uma pasta separada
            if 0 < len(cnpj_cliente) < 14:
                escreve_relatorio_csv(f"NFSe_{numero_nota}.pdf;Tomador com CPF, não gera arquivo para importar;'{cnpj_cliente}", nome='Andamentos', local=pasta_final)
                pasta_final_nota = path.join(pasta_final, 'Notas CPF')
                nome_nota_final = f'nfe_{numero_nota}_tomador_{cnpj_cliente}.pdf'
                makedirs(pasta_final_nota, exist_ok=True)
                
            # se não constar CNPJ do cliente na nota, anota a ocorrência
            elif cnpj_cliente == '':
                escreve_relatorio_csv(f"NFSe_{numero_nota}.pdf;Tomador sem CNPJ informado", nome='Andamentos', local=pasta_final)
            
            # se tiver busca o código do Domínio na planilha
            else:
                codigo_cliente = busca_codigo_cliente(cnpj_cliente)
                # se não encontrar o CNPJ do cliente na planilha de códigos, anota a ocorrência
                if codigo_cliente == 'Cliente não encontrado na planilha de códigos':
                    escreve_relatorio_csv(f"NFSe_{numero_nota}.pdf;{codigo_cliente};'{cnpj_cliente}", nome='Andamentos', local=pasta_final)
                elif codigo_cliente == 'CÓDIGO DO DOMÍNIO':
                    escreve_relatorio_csv(f"NFSe_{numero_nota}.pdf;CNPJ do cliente repetido na planilha de códigos, verificar qual consta com o código correto e deletar os demais;'{cnpj_cliente}", nome='Andamentos', local=pasta_final)
                # se o CNPJ do cliente for encontrado na planilha de códigos, cria arquivo .txt para importar no dóminio web
                else:
                    cria_txt(codigo_cliente, cnpj_cliente, f'{dados_notas_escritorio};{dados_notas}', pasta_final)
                    escreve_relatorio_csv(f"NFSe_{numero_nota}.pdf;Arquivo para importação criado com sucesso;{codigo_cliente} - {cnpj_cliente}", nome='Andamentos', local=pasta_final)
                    
            escreve_relatorio_csv(f'{dados_notas_escritorio};{dados_notas_cliente};{dados_notas}', nome='Dados das notas', local=pasta_final)

    # renomeia a nota colocando o número dela
    move(arq, path.join(pasta_final_nota, nome_nota_final))
    return


def cria_txt(codigo_cliente, cnpj_cliente, dados_nota, pasta_final, tipo='NFSe', encode='latin-1'):
    # cria um txt com os dados das notas em uma pasta nomeada com o código no domínio e o cnpj do prestador
    local = path.join(pasta_final, 'Arquivos para Importação', str(codigo_cliente) + '-' + str(cnpj_cliente))
    makedirs(local, exist_ok=True)
    
    try:
        f = open(path.join(local, f"{tipo} - {cnpj_cliente}.txt"), 'a', encoding=encode)
    except:
        f = open(path.join(local, f"{tipo}  - {cnpj_cliente} - auxiliar.txt"), 'a', encoding=encode)
    
    f.write(str(dados_nota) + '\n')
    f.close()
    

def busca_codigo_cliente(cnpj_cliente):
    try:
        # Localiza a planilha
        caminho = path.join('Dados_clientes', 'códigos clientes.xlsx')
        # Abrir a planilha
        lista = read_excel(caminho)
        # Definir o index da planilha
        lista.set_index('CNPJ', inplace=True)
        
        cliente = lista.loc[int(cnpj_cliente)]
        cliente = list(cliente)
        codigo_tomador = cliente[0]
        
        return codigo_tomador
    except:
        return 'Cliente não encontrado na planilha de códigos'
    

def run(window, input_excel, output_dir):
    arquivo, notas, tipo_dados = open_lista_dados(input_excel)
    pasta_final = output_dir
    # Abre o site do SIGISS pronto para validar as notas
    with Session() as s:
        soup = ''
        timer = 0
        timer_2 = -1
        # enquanto o site não carregar, tenta acessá-lo e caso demore muito retorna
        while not compile(r'<input class=\"\" id=\"codigo\" label=\"Informe o Código de Autenticidade : \" maxlength=\"42\" name=\"codigo\" requerido=\"true\" type=\"text\"/>').search(soup):
            sleep(0.99)
            try:
                response = s.get('https://valinhos.sigissweb.com/validarnfe')
                soup = BeautifulSoup(response.content, 'html.parser', from_encoding="Latin-1")
                soup = soup.prettify()
            except:
                soup = ''
                pass
            
            timer += 1
            if timer >= 2:
                timer_2 += 1
                # exibe um contador de espera na interface
                window['-Mensagens-'].update('Aguardando SIGISSWEB... ' + str(timer_2))
                
            if timer >= 32:
                window['-Mensagens-'].update('')
                alert(text='SIGISSWEB demorou muito para responder, download encerrado.\n\n'
                           'O site pode estar fora do ar ou foi movido para outro endereço.\n\n'
                           'Verifique se o site "https://valinhos.sigissweb.com/validarnfe" ainda existe.\n\n')
                
                s.close()
                return
            
            # Verifica se o usuário solicitou o encerramento do script
            if event == '-Encerrar-':
                s.close()
                return
        
        # quando o site abrir
        # apaga qualquer mensagem na interface
        window['-Mensagens-'].update('')
        # configura a pasta final dos arquivos
        pasta_final = path.join(pasta_final, 'Notas Fiscais de Serviço')
        # inicia a variável do contador de notas baixadas
        quantidade_notas = 0
        # atualiza a barra de progresso para ela ficar mais visível
        window['-progressbar-'].update(bar_color=('#fca400', '#ffe0a6'))
        
        if tipo_dados == 'xls':
            # cria o indice para cada empresa da lista de dados
            total_empresas = range(notas.nrows)
            for count, nota in enumerate(range(notas.nrows), start=1):
                # pega a última coluna da planilha
                nota = notas.cell_value(nota, -1)
                if download_nota(nota, s, pasta_final):
                    quantidade_notas += 1
                # atualiza a barra de progresso
                window['-progressbar-'].update_bar(count, max=int(len(total_empresas)))
                window['-Progresso_texto-'].update(str( round( float(count) / int(len(total_empresas)) *100, 1) ) + '%')
                window.refresh()
                
                # Verifica se o usuário solicitou o encerramento do script
                if event == '-Encerrar-':
                    s.close()
                    return
                
        elif tipo_dados == 'xlsx':
            last_column = notas.max_column
            # cria o indice para cada empresa da lista de dados
            total_empresas = []
            for nota in notas.iter_rows(min_row=1, min_col=last_column, max_col=last_column):
                total_empresas.append(nota)
            
            for count, nota in enumerate(notas.iter_rows(min_row=1, min_col=last_column, max_col=last_column), start=1):
                for chave in nota:
                    if download_nota(chave.value, s, pasta_final):
                        quantidade_notas += 1
                    # atualiza a barra de progresso
                    window['-progressbar-'].update_bar(count, max=int(len(total_empresas)))
                    window['-Progresso_texto-'].update(str( round( float(count) / int(len(total_empresas)) *100, 1) ) + '%')
                    window.refresh()

                    # Verifica se o usuário solicitou o encerramento do script
                    if event == '-Encerrar-':
                        s.close()
                        return
    
        s.close()
        alert(text=f'Download concluído, {quantidade_notas} notas salvas')
        # ao terminar a execução, apaga a porcentagem e a barra de progresso para a interface ficar mais limpa
        window['-progressbar-'].update_bar(0)
        window['-progressbar-'].update(bar_color='#f0f0f0')
        window['-Progresso_texto-'].update('')
        window['-Mensagens-'].update('')
        
    

# Define o ícone global da aplicação
sg.set_global_icon('Assets/auto-flash.ico')
if __name__ == '__main__':
    sg.theme('GrayGrayGray')  # Define o tema do PySimpleGUI
    # sg.theme_previewer()
    # Layout da janela
    layout = [
        [sg.Button('Ajuda', border_width=0), sg.Button('Lista de códigos do Domínio', border_width=0), sg.Button('Log do sistema', border_width=0, disabled=True)],
        [sg.Text('')],
        [sg.Text('Selecione um arquivo Excel com as chaves de acesso das notas:')],
        [sg.FileBrowse('Pesquisar', key='-Abrir-', file_types=(('Planilhas Excel', '*.xlsx *.xls'),)), sg.InputText(key='-input_excel-', size=80, disabled=True)],
        [sg.Text('Selecione um diretório para salvar os resultados:')],
        [sg.FolderBrowse('Pesquisar', key='-Abrir2-'), sg.InputText(key='-output_dir-', size=80, disabled=True)],
        [sg.Text('')],
        [sg.Text('', key='-Mensagens-')],
        [sg.Text(size=6, text='', key='-Progresso_texto-'), sg.ProgressBar(max_value=0, orientation='h', size=(54,5), key='-progressbar-', bar_color='#f0f0f0')],
        [sg.Button('Iniciar', key='-Iniciar-', border_width=0), sg.Button('Encerrar', key='-Encerrar-', disabled=True, border_width=0), sg.Button('Abrir resultados', key='-Abrir resultados-', disabled=True, border_width=0)],
    ]
    
    # guarda a janela na variável para manipula-la
    window = sg.Window('Download NFSE_VP SIGISSWEB', layout)
    
    def run_script_thread():
        try:
            if not input_excel or not output_dir:
                alert(text=f'Por favor selecione uma planilha e um diretório.')
                return
            
            # habilita e desabilita os botões conforme necessário
            window['-Abrir-'].update(disabled=True)
            window['-Abrir2-'].update(disabled=True)
            window['-Iniciar-'].update(disabled=True)
            window['-Encerrar-'].update(disabled=False)
            window['-Abrir resultados-'].update(disabled=False)
            
            try:
                # Chama a função que executa o script
                run(window, input_excel, output_dir)
            # Qualquer erro o script exibe um alerta e salva gera o arquivo log de erro
            except Exception as erro:
                window['Log do sistema'].update(disabled=False)
                alert(text=f"Erro :'(\n\n"
                           f'Abra o pasta de "Log do sistema" e envie o arquivo "Log.txt" para o desenvolvedor.\n')
                escreve_doc(erro)
            
            # habilita e desabilita os botões conforme necessário
            window['-Abrir-'].update(disabled=False)
            window['-Abrir2-'].update(disabled=False)
            window['-Iniciar-'].update(disabled=False)
            window['-Encerrar-'].update(disabled=True)
            # apaga a porcentagem e a barra de progresso para a interface ficar mais limpa
            window['-progressbar-'].update_bar(0)
            window['-progressbar-'].update(bar_color='#f0f0f0')
            window['-Progresso_texto-'].update('')
            window['-Mensagens-'].update('')
        except:
            pass
    
    
    while True:
        # captura o evento e os valores armazenados na interface
        event, values = window.read()
        try:
            input_excel = values['-input_excel-']
            output_dir = values['-output_dir-']
        except:
            input_excel = 'Desktop'
            output_dir = 'Desktop'
        
        if event == sg.WIN_CLOSED:
            break
        
        elif event == 'Log do sistema':
            startfile('Log')
            
        elif event == 'Ajuda':
            startfile('Manual do usuário - Download NFSe_VP SIGISSWEB.pdf')
          
        elif event == 'Lista de códigos do Domínio':
            startfile(path.join('Dados_clientes', 'códigos clientes.xlsx'))
            
        elif event == '-Iniciar-':
            # Cria uma nova thread para executar o script
            script_thread = Thread(target=run_script_thread)
            script_thread.start()
            
        elif event == '-Encerrar-':
            window['-Encerrar-'].update(disabled=True)
            window['-Iniciar-'].update(disabled=False)
            window['-Abrir resultados-'].update(disabled=False)
            alert(text='Download encerrado.\n\n'
                     'Caso queira reiniciar o download, apague os arquivos gerados anteriormente ou selecione um novo local.\n\n'
                     'O Script não continua uma execução já iniciada.\n\n')
            
        elif event == '-Abrir resultados-':
            startfile(path.join(output_dir, 'Notas Fiscais de Serviço'))
        
    window.close()
